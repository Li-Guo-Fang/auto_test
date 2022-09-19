import os
from openpyxl import load_workbook
from settings.config import BASE_DIR

# workbook = load_workbook(os.path.join(BASE_DIR,'data','testcase.xlsx'))
# sheet = workbook.worksheets[0]
#
# print(sheet.max_column)
# print(sheet.max_row)
# print(sheet.cell(2,2).value)
#
# for row in sheet.rows:
#     for cell in row:
#         print(cell.value)
#
# print('-'*100)
#
# for col in sheet.columns:
#     for cell in col:
#         print(cell.value)



class ExcelHandler:
    def __init__(self,file_name):
        self.file_name = os.path.join(BASE_DIR,'data',file_name)


    def handler(self):
        workbook = load_workbook(self.file_name)
        sheet = workbook.worksheets[0]
        return sheet

    def read_rows(self):
        datas_rows_list = []
        for row in self.handler().rows:
            row_list = []
            for cell in row:
                row_list.append(cell.value)
            datas_rows_list.append(row_list)
        return datas_rows_list

    def read_cols(self):
        datas_col_list = []
        for col in self.handler().columns:
            col_list = []
            for cell in col:
                col_list.append(cell.value)
            datas_col_list.append(col_list)
        return datas_col_list


if __name__ == '__main__':
    excel = ExcelHandler('testcase.xlsx')
    print(excel.read_rows())
    print('-'*100)
    print(excel.read_cols())



